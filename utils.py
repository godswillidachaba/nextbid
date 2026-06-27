import logging
import re
import time
from datetime import datetime, timezone
from functools import wraps
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

_ILLEGAL_CHARS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

def _sanitize(value):
    """Remove characters illegal in openpyxl worksheet cells."""
    if isinstance(value, str):
        return _ILLEGAL_CHARS_RE.sub('', value)
    return value


# ── Excel style constants ─────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(bold=True, size=14, color='1a1a2e')
NIGERIA_FILL = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
LIBERIA_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
SCORE_HIGH_FONT = Font(bold=True, color='10b981')
SCORE_MED_FONT = Font(bold=True, color='f59e0b')
SCORE_LOW_FONT = Font(bold=True, color='ef4444')
DATA_HEADERS = [
    'ID', 'Notice ID', 'Title', 'Description', 'Organization',
    'Country', 'Deadline', 'Published', 'URL', 'Source',
    'Score', 'Position', 'Unit', 'Executive Summary',
    'Scraped At', 'Analyzed At',
]

TARGET_COUNTRIES = ['nigeria', 'liberia']

NIGERIA_CITIES = [
    'lagos', 'abuja', 'kano', 'ibadan', 'kaduna', 'port harcourt', 'benin city',
    'maiduguri', 'zaria', 'aba', 'jos', 'ilorin', 'oyo', 'enugu', 'abeokuta',
    'onitsha', 'warri', 'sokoto', 'katsina', 'bauchi', 'akure', 'calabar',
    'owerri', 'uyo', 'iffo', 'adikpo', 'gashaka', 'gembu', 'awka', 'ekiti',
    'osogbo', 'yola', 'makurdi', 'minna', 'gombe', 'jalingo', 'damaturu',
    'kebbi', 'lafia', 'birnin kebbi', 'katsina-ala', 'nguru', 'potiskum',
]

LIBERIA_CITIES = [
    'monrovia', 'bensonville', 'gbarnga', 'harbel', 'kakata',
    'zwedru', 'tubmanburg', 'buchanan', 'harper',
]


def retry(max_attempts=3, delay=5):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            last_exc = None
            for attempt in range(1, max_attempts + 1):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    last_exc = e
                    if attempt < max_attempts:
                        logger.warning('%s attempt %d/%d failed: %s — retrying in %ds',
                                       func.__name__, attempt, max_attempts, e, delay)
                        time.sleep(delay)
            logger.error('%s failed after %d attempts: %s', func.__name__, max_attempts, last_exc)
            raise last_exc
        return wrapper
    return decorator


def matches_target_country(text):
    if not text:
        return False
    lower = text.lower()
    if any(c in lower for c in TARGET_COUNTRIES):
        return True
    if any(c in lower for c in NIGERIA_CITIES):
        return True
    if any(c in lower for c in LIBERIA_CITIES):
        return True
    return False


def _get_country_fill(country):
    """Return Nigeria or Liberia fill, or None."""
    if not country:
        return None
    lower = country.lower()
    if 'nigeria' in lower:
        return NIGERIA_FILL
    if 'liberia' in lower:
        return LIBERIA_FILL
    return None


def _write_section_header(ws, row, text, max_col=2):
    """Write a merged section header with styling."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')


def _build_summary_stats(notices, analyses):
    """Compute all summary statistics from notices + analyses."""
    now = datetime.utcnow()
    current_month = now.month
    current_year = now.year

    total = len(notices)
    analyzed_count = sum(1 for n in notices if n.notice_id in analyses)
    unanalyzed = total - analyzed_count

    nigeria = sum(1 for n in notices if n.country and 'nigeria' in n.country.lower())
    liberia = sum(1 for n in notices if n.country and 'liberia' in n.country.lower())
    other = total - nigeria - liberia

    deadlines_this_month = 0
    deadlines_past = 0
    no_deadline = 0
    for n in notices:
        if n.deadline_date is None:
            no_deadline += 1
        else:
            dt = n.deadline_date
            if dt.month == current_month and dt.year == current_year:
                deadlines_this_month += 1
            if dt < now:
                deadlines_past += 1

    high = sum(1 for n in notices if analyses.get(n.notice_id) and analyses[n.notice_id].score is not None and analyses[n.notice_id].score >= 70)
    med = sum(1 for n in notices if analyses.get(n.notice_id) and analyses[n.notice_id].score is not None and 40 <= analyses[n.notice_id].score < 70)
    low = sum(1 for n in notices if analyses.get(n.notice_id) and analyses[n.notice_id].score is not None and analyses[n.notice_id].score < 40)

    scores = [analyses[n.notice_id].score for n in notices if analyses.get(n.notice_id) and analyses[n.notice_id].score is not None]
    avg_score = round(sum(scores) / len(scores), 1) if scores else 0

    source_counts = {}
    for n in notices:
        s = n.source or 'Unknown'
        source_counts[s] = source_counts.get(s, 0) + 1

    scored = [(n, analyses[n.notice_id]) for n in notices if n.notice_id in analyses and analyses[n.notice_id].score is not None]
    scored.sort(key=lambda x: x[1].score, reverse=True)
    top5 = scored[:5]

    return {
        'total': total,
        'analyzed': analyzed_count,
        'unanalyzed': unanalyzed,
        'nigeria': nigeria,
        'liberia': liberia,
        'other': other,
        'deadlines_this_month': deadlines_this_month,
        'deadlines_past': deadlines_past,
        'no_deadline': no_deadline,
        'high': high,
        'med': med,
        'low': low,
        'avg_score': avg_score,
        'source_counts': source_counts,
        'top5': top5,
    }


def _write_data_sheet(ws, notices, analyses, headers=None):
    """Write a styled data sheet sorted by score desc with country-based coloring."""
    if headers is None:
        headers = DATA_HEADERS

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    def sort_key(n):
        ba = analyses.get(n.notice_id)
        if ba and ba.score is not None:
            return -ba.score
        return float('-inf')

    sorted_notices = sorted(notices, key=sort_key)

    for row_idx, n in enumerate(sorted_notices, 2):
        ba = analyses.get(n.notice_id)
        fill = _get_country_fill(n.country)

        row_data = [
            n.id, n.notice_id,
            _sanitize(n.title), _sanitize(n.description),
            _sanitize(n.organization),
            _sanitize(n.country),
            _sanitize(n.deadline), _sanitize(n.published),
            _sanitize(n.url), _sanitize(n.source),
            ba.score if ba else '',
            _sanitize(ba.suggested_position) if ba else '',
            _sanitize(ba.relevant_unit) if ba else '',
            _sanitize(ba.executive_summary) if ba else '',
            n.scraped_at.isoformat() if n.scraped_at else '',
            ba.analyzed_at.isoformat() if ba and ba.analyzed_at else '',
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if fill:
                cell.fill = fill

    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def generate_excel_bytes(notices, analyses):
    """Generate a styled Excel workbook with Summary dashboard + All Data + per-source sheets.

    Args:
        notices: List of objects with .id, .notice_id, .title, .organization,
                 .country, .deadline, .published, .url, .source, .scraped_at
        analyses: Dict mapping notice_id -> object with .score, .suggested_position,
                  .relevant_unit, .analyzed_at

    Returns:
        BytesIO of the .xlsx file
    """
    wb = Workbook()
    stats = _build_summary_stats(notices, analyses)
    now = datetime.now(timezone.utc)

    # ── Summary Sheet (stats dashboard) ──
    ws = wb.active
    ws.title = 'Summary'
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 15

    ws.cell(row=1, column=1, value='Nextier Opportunity Finder — Summary Report').font = TITLE_FONT
    ws.cell(row=3, column=1, value=f'Generated: {now.strftime("%Y-%m-%d %H:%M UTC")}')

    _write_section_header(ws, 5, 'DESCRIPTION')
    ws.cell(row=6, column=1, value='This report summarizes procurement opportunities scraped from 8 sources, targeting Nigeria and Liberia.').font = Font(italic=True, color='4a4a6a')

    _write_section_header(ws, 8, 'SCRAPING OVERVIEW')
    ws.cell(row=9, column=1, value='Total Notices Scraped')
    ws.cell(row=9, column=2, value=stats['total'])
    ws.cell(row=10, column=1, value='Notices with AI Analysis')
    ws.cell(row=10, column=2, value=stats['analyzed'])
    ws.cell(row=11, column=1, value='Unanalyzed Notices')
    ws.cell(row=11, column=2, value=stats['unanalyzed'])

    _write_section_header(ws, 13, 'COUNTRY BREAKDOWN')
    ws.cell(row=14, column=1, value='Nigeria')
    c = ws.cell(row=14, column=2, value=stats['nigeria'])
    c.fill = NIGERIA_FILL
    ws.cell(row=15, column=1, value='Liberia')
    c = ws.cell(row=15, column=2, value=stats['liberia'])
    c.fill = LIBERIA_FILL
    ws.cell(row=16, column=1, value='Other / Multiple')
    ws.cell(row=16, column=2, value=stats['other'])

    _write_section_header(ws, 18, 'DEADLINES')
    ws.cell(row=19, column=1, value=f'Deadlines This Month ({now.strftime("%B %Y")})')
    ws.cell(row=19, column=2, value=stats['deadlines_this_month'])
    ws.cell(row=20, column=1, value='Past Due Deadlines')
    ws.cell(row=20, column=2, value=stats['deadlines_past'])
    ws.cell(row=21, column=1, value='No Deadline Set')
    ws.cell(row=21, column=2, value=stats['no_deadline'])

    _write_section_header(ws, 23, 'SCORE DISTRIBUTION')
    ws.cell(row=24, column=1, value='High (Score 70+)')
    ws.cell(row=24, column=2, value=stats['high']).font = SCORE_HIGH_FONT
    ws.cell(row=25, column=1, value='Medium (Score 40-69)')
    ws.cell(row=25, column=2, value=stats['med']).font = SCORE_MED_FONT
    ws.cell(row=26, column=1, value='Low (Score <40)')
    ws.cell(row=26, column=2, value=stats['low']).font = SCORE_LOW_FONT
    ws.cell(row=27, column=1, value='Average Score')
    ws.cell(row=27, column=2, value=stats['avg_score'])

    src_start = 29
    _write_section_header(ws, src_start, 'SOURCE BREAKDOWN')
    for i, (src, count) in enumerate(sorted(stats['source_counts'].items()), 1):
        ws.cell(row=src_start + i, column=1, value=src)
        ws.cell(row=src_start + i, column=2, value=count)

    top5_start = src_start + len(stats['source_counts']) + 2
    ws.merge_cells(start_row=top5_start, start_column=1, end_row=top5_start, end_column=5)
    c = ws.cell(row=top5_start, column=1, value='TOP 5 OPPORTUNITIES')
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal='center')
    for col in range(1, 6):
        ws.cell(row=top5_start, column=col).fill = HEADER_FILL

    top5_headers = ['Title', 'Score', 'Position', 'Country', 'Deadline']
    hr = top5_start + 1
    for ci, h in enumerate(top5_headers, 1):
        ws.cell(row=hr, column=ci, value=h).font = Font(bold=True)

    for i, (n, ba) in enumerate(stats['top5']):
        r = hr + 1 + i
        fill = _get_country_fill(n.country)
        title = _sanitize((n.title or '')[:80])
        url = _sanitize(n.url or '')
        if url:
            safe_url = url.replace('"', '""')
            safe_title = title.replace('"', '""')
            cell = ws.cell(row=r, column=1)
            cell.value = f'=HYPERLINK("{safe_url}", "{safe_title}")'
            cell.font = Font(color='0563C1', underline='single')
        else:
            ws.cell(row=r, column=1, value=title)
        for ci, v in enumerate([
            ba.score,
            _sanitize(ba.suggested_position or ''),
            _sanitize(n.country or ''),
            _sanitize(n.deadline or ''),
        ], 2):
            cell = ws.cell(row=r, column=ci, value=v)
            if fill:
                cell.fill = fill

    # ── All Data Sheet ──
    ws_all = wb.create_sheet(title='All Data')
    _write_data_sheet(ws_all, notices, analyses)

    # ── Per-Source Sheets ──
    sources = sorted(set(n.source for n in notices))
    for src in sources:
        src_notices = [n for n in notices if n.source == src]
        ws_src = wb.create_sheet(title=src[:31])
        _write_data_sheet(ws_src, src_notices, analyses)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
